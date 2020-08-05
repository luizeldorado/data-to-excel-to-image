# Warning: shit coding ahead

make_image_files = True
remove_excel_files = True

data_filename = "dados.xlsx"
template_filename = "modelo.xlsx"
excel_prefix = "temp "
image_prefix = "boletim "

###########

'''
TODO:
- Search for tags only once in template, instead of each time. Like, save the cells and positions of tags and just use them later. (would still have to reload it to reset to original template content)
- Remove excel2img, since it reopens the file. Make it so we use openpyxl, maybe in conjunction with win32 API.
- Save and restore clipboard state. Yep, it actually puts stuff in there to save the image. I don't think that there is a function on the API to save the image directly.
- Make those options above into command line args
- Maybe make a GUI
- Be able to select multiple worksheets in the template file, each one would make one file
- Fix/add other things that I can't remember right now

'''

import excel2img
import openpyxl
import os
from collections import namedtuple
import logging
from logging import debug, info, warning, error, critical

logging.basicConfig(format='%(levelname)s: %(message)s', level=logging.INFO)

stop = False

if not os.path.isfile(data_filename):
	critical("Data file '{}' does not exist!".format(data_filename))
	stop = True
if not os.path.isfile(template_filename):
	critical("Template file '{}' does not exist!".format(template_filename))
	stop = True

if stop:
	exit()

debug("Opening data file '{}'".format(data_filename))

wb = openpyxl.load_workbook(data_filename)
ws = wb.active

debug("Getting field names:")

fields = []
for cell in ws[1]:
	fields.append(str(cell.value))

debug(fields)

debug("Getting each data row:")

for data_row in ws.iter_rows(min_row=2):

	info("Generating row '{}'".format(str(data_row[0].value)))

	debug("- (Re)opening template file '{}'".format(template_filename))

	wb = openpyxl.load_workbook(template_filename)
	ws = wb.active

	debug("- Looping through all non-empty cells to find tags:")

	for template_row in ws.rows:
		for cell in template_row:
			if cell.value != None:

				# find tags in that cell

				cell_str = str(cell.value)
				pos = 0

				debug("  - Cell {}: {}".format(cell.coordinate, cell_str))

				while True:

					# find next tag after pos
					percent_pos_start = cell_str.find("%", pos)

					if percent_pos_start != -1:
						# find end of tag
						percent_pos_end = cell_str.find("%", percent_pos_start+1)

						if percent_pos_end == -1:
							debug("    - percent_pos_start = {}, no ending percent".format(percent_pos_start))
							break  # no ending %, so there is no more tags (no point in warning, probably an actual percentage sign there

						# find name of tag
						tag_name = cell_str[percent_pos_start+1:percent_pos_end]

						debug("    - percent_pos_start = {}, percent_pos_end = {}, tag_name = {}".format(percent_pos_start, percent_pos_end, tag_name))

						# find which field is this tag
						for field_i, field in enumerate(fields):

							if tag_name == field:

								# replace tag name and %s with value in data

								data_value = data_row[field_i].value
								data_type = data_row[field_i].data_type

								debug("      - data_value = {}, data_type = {}".format(repr(data_value), data_type), )

								data_value_str = str(data_value) if data_value != None else ""

								# check if the cell only contains the tag, if so copy the type directly

								if percent_pos_start == 0 and percent_pos_end+1 == len(cell_str):

									cell_str = data_value_str
									cell.value = data_value
									# cell.data_type = data_type

									debug("        - whole tag cell copied directly, cell.data_type = {}".format(cell.data_type))

								else:

									cell_str = cell_str[:percent_pos_start] + data_value_str + cell_str[percent_pos_end+1:]
									cell.value = cell_str

								pos = percent_pos_start + len(data_value_str)

								debug("      - replaced, cell_str = {}".format(cell_str))

								break

						else:
							debug("      - no field called {}".format(tag_name))
							warning("%{}% found in template file, but not in the data file. Did you mispell it or didn't add it to the template?".format(tag_name))
							pos = percent_pos_end+1

						debug("      - pos = {}".format(pos))

					else:
						# no more %s
						break

	excel_filename = excel_prefix + str(data_row[0].value) + ".xlsx"
	image_filename = image_prefix + str(data_row[0].value) + ".png"

	debug("- Saving excel file {}".format(excel_filename))

	wb.save(excel_filename)

	if make_image_files:
		debug("- Saving image file {}".format(image_filename))
		excel2img.export_img(excel_filename, image_filename, None, None)

	if remove_excel_files:
		debug("- Removing excel file {}".format(excel_filename))
		os.remove(excel_filename)