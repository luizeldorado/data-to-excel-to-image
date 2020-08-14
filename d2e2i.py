# d2e2i

from collections import namedtuple
import copy
import excel2img
import os
import openpyxl
import pathlib
from pprint import pprint

class D2E2I():

	def __init__(self):

		self.file_template = ""
		self.file_data = ""
		self.generate_excel_files = False
		self.generate_image_files = False
		self.prefix = ""
		self.folder_excel = ""
		self.folder_image = ""
		self.mark_start = "%"
		self.mark_end = "%"

		self.wb_template = None
		self.wb_data = None
		self.ws_template = None
		self.ws_data = None

	def validate_file_template(self):
		return os.path.isfile(self.file_template)
	def validate_file_data(self):
		return os.path.isfile(self.file_data)

	# Open both template and data file. Must be closed afterwards
	def open_files(self):
		if not self.validate_file_template() or not self.validate_file_data():
			return False

		# Close data file if open
		self.close_files()

		# Open data file
		self.wb_data = openpyxl.load_workbook(self.file_data, read_only=True)
		self.ws_data = self.wb_data.active

		# Open template file
		self.wb_template = openpyxl.load_workbook(self.file_template)
		self.ws_template = self.wb_template.active

		return True

	def close_files(self):
		# Apparently I have to close this because it's read only... TODO: find out about this
		if self.wb_data:
			self.wb_data.close()

	def ws_iter_cells(self, ws):
		for row in ws.rows:
			for cell in row:
				yield cell

	def get_tag_cells_in_template(self):

		tag_cells = []

		for template_cell in self.ws_iter_cells(self.ws_template):

			if template_cell.value == None:
				continue

			tags = []
			cell_string = str(template_cell.value)
			
			pos = 0

			while True:
				# Find next starting %
				pos_tag_start = cell_string.find("%", pos)
				if pos_tag_start == -1:
					break
				pos = pos_tag_start + 1

				# Find next ending %
				pos_tag_end = cell_string.find("%", pos)
				if pos_tag_end == -1:
					break
				
				# Check if field name actually exists in data
				tag_field_name = cell_string[pos_tag_start+1:pos_tag_end]
				tag_field_column = 0

				for field_cell in self.ws_data[1]:
					if str(field_cell.value) == tag_field_name:
						tag_field_column = field_cell.column

				if tag_field_column != 0:

					if pos_tag_start == 0 and pos_tag_end == len(cell_string)-1:
						# Tag fills the whole cell, no position
						tag_pos = None
					else:
						tag_pos = TagPos(pos_tag_start, pos_tag_end)

					tags.append(Tag(
						field_name=tag_field_name,
						field_column=tag_field_column,
						pos=tag_pos
					))

					pos = pos_tag_end + 1

				else:
					# Start % is ignored but end % could be next start %
					pos = pos_tag_end

			if len(tags) != 0:
				tag_cells.append(
					TagCell(template_cell.row, template_cell.column, tags)
				)

		return tag_cells

	# Yields each row of data file, ignoring the first (that's the fields row).
	def iter_data_rows(self):
		for data_row in self.ws_data.iter_rows(min_row=2):
			yield data_row

	def number_of_data_rows(self):
		return self.ws_data.max_row - 1

	def reload_template(self):
		# TODO: I think this reads the file from disk again... maybe make a copy in memory?
		self.wb_template = openpyxl.load_workbook(self.file_template)
		self.ws_template = self.wb_template.active

	def generate_row(self, data_row, tag_cells):

		# Reload/reopen template file
		self.reload_template()

		# Replace all tags on cells that have them
		self.replace_tag_cells_in_template(tag_cells, data_row)

		# Save template file
		filename_excel = os.path.join(self.folder_excel, self.prefix + str(data_row[0].value) + ".xlsx")
		self.wb_template.save(filename_excel)

		# Save image file
		if self.generate_image_files:
			filename_image = os.path.join(self.folder_image, self.prefix + str(data_row[0].value) + ".png")
			excel2img.export_img(filename_excel, filename_image, None, None)

		if not self.generate_excel_files:
			os.remove(filename_excel)

	def replace_tag_cells_in_template(self, tag_cells, data_row):

		for tag_cell in tag_cells:
				
			cell = self.ws_template.cell(tag_cell.row, tag_cell.column)
			pos_offset = 0

			for tag in tag_cell.tags:

				data_field_value = data_row[tag.field_column-1].value

				if tag.pos:
					# Tag has position inside text
					cell_value_str = str(cell.value)

					if data_field_value == None:
						data_field_value_str = ""
					else:
						data_field_value_str = str(data_field_value)

					cell.value = (
						cell_value_str[:pos_offset+tag.pos.start] +
						data_field_value_str +
						cell_value_str[pos_offset+tag.pos.end+1:]
					)

					pos_offset += len(data_field_value_str) - (tag.pos.end+1 - tag.pos.start)

				else:
					# Tag fills the whole cell
					cell.value = data_field_value

	def generate(self):
		
		# Open wb_template and wb_data.
		if not self.open_files():
			return False

		# Find tags inside template
		tag_cells = self.get_tag_cells_in_template()
		print("tag_cells")
		pprint(tag_cells)

		# Make results folder
		pathlib.Path(self.folder_excel).mkdir(parents=True, exist_ok=True)

		# Read each row of data file
		for data_row in self.iter_data_rows():
			self.generate_row(data_row, tag_cells)
			
		self.close_files()

		return True

TagCell = namedtuple('TagCell', ['row', 'column', 'tags'])
Tag = namedtuple('Tag', ['field_name', 'field_column', 'pos'])
TagPos = namedtuple('TagPos', ['start', 'end'])

# if __name__ == "__main__":
# 	pass

'''
TODO:
- Remove excel2img, since it reopens the file. Make it so we use openpyxl, maybe in conjunction with win32 API.
- Save and restore clipboard state. Yep, it actually puts stuff in there to save the image. I don't think that there is a function on the API to save the image directly.
- Make those options above into command line args
- Be able to select multiple worksheets in the template file, each one would make one file
- Fix/add other things that I can't remember right now
'''